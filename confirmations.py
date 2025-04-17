import ftplib
import json
import logging
import shutil
import time
import openpyxl
from datetime import datetime
from pprint import pprint
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.worksheet import Worksheet

import pandas as pd
import requests

from settings import *

logging.basicConfig(
    filename=f"{BASE_DIR}/confirmations.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)

def download_from_ftp(server, username, password, local_directory, division_id, max_retries=3, retry_delay=5):
    remote_directory = "/OZON_orders/Frenchpharmacy/Confirmations"
    attempt = 0

    while attempt < max_retries:
        try:
            with ftplib.FTP(server) as ftp:
                try:
                    ftp.login(user=username, passwd=password)
                    logging.info("Успешное подключение и авторизация на сервере")
                except ftplib.error_perm as e:
                    logging.error(f"Ошибка авторизации: {e}. Попытка {attempt + 1} из {max_retries}")
                    time.sleep(retry_delay)
                    attempt += 1
                    continue

                try:
                    ftp.cwd(remote_directory)
                    logging.info("Успешный переход в директорию на сервере")
                except ftplib.error_perm as e:
                    logging.error(f"Ошибка при переходе в директорию: {e}. Попытка {attempt + 1} из {max_retries}")
                    time.sleep(retry_delay)
                    attempt += 1
                    continue

                if not os.path.exists(local_directory):
                    os.makedirs(local_directory)

                files = ftp.nlst()
                found_any = False

                for filename in files:
                    if (
                        filename.endswith('.xls') and
                        '_' in filename and
                        filename.rsplit('_', 1)[-1].replace('.xls', '') == division_id
                    ):
                        found_any = True
                        order_number = filename.rsplit('_', 1)[0]
                        new_filename = f"{order_number}.xls"
                        local_path = os.path.join(local_directory, new_filename)

                        try:
                            with open(local_path, 'wb') as local_file:
                                ftp.retrbinary(f"RETR {filename}", local_file.write)
                                logging.info(f"Файл скачан и переименован: {filename} → {new_filename}")

                            ftp.delete(filename)
                            logging.info(f"Файл удален с сервера: {filename}")

                        except (ftplib.all_errors, ConnectionError) as e:
                            logging.error(f"Ошибка при скачивании {filename}: {e}. Попытка {attempt + 1} из {max_retries}")
                            time.sleep(retry_delay)

                if not found_any:
                    logging.info(f"Файлы с кодом подразделения {division_id} не найдены в директории.")

                return  # Завершаем после успешной попытки

        except (ftplib.all_errors, ConnectionError) as e:
            logging.error(f"Ошибка соединения с FTP-сервером: {e}. Попытка {attempt + 1} из {max_retries}")
            time.sleep(retry_delay)

        attempt += 1

    logging.error("Не удалось подключиться к FTP-серверу после нескольких попыток")


def append_data_from_file(file_path, df=None):
    """
    Читает данные из Excel-файла и добавляет их в DataFrame.
    Если DataFrame не передан, создается новый.

    :param file_path: str - Путь к файлу Excel.
    :param df: pd.DataFrame - Существующий DataFrame (по умолчанию None).
    :return: pd.DataFrame - Обновленный DataFrame.
    """
    try:
        # Чтение данных из Excel в DataFrame
        file_data = pd.read_excel(file_path, dtype={'CODEART': str, 'CODEPST': str})

        # Создание нового DataFrame, если не передан существующий
        if df is None:
            df = file_data
        else:
            # Объединение данных
            df = pd.concat([df, file_data], ignore_index=True)

        df = df.fillna('нет данных')
        return df

    except Exception as e:
        logging.error(f"Ошибка при чтении файла при создании датафрейма {file_path}: {e}")
        return df



def validate_dataframe(df):
    """
    Проверка наличия колонки HDRTAG2
    """
    if "HDRTAG2" not in df.columns:
        logging.error("Колонка 'HDRTAG2' не найдена в DataFrame.")
        return False
    return True


def extract_unique_orders(df):
    """
    Удаление дубликатов по HDRTAG2
    """
    try:
        return df.drop_duplicates(subset=["HDRTAG2"])
    except Exception as e:
        logging.error(f"Ошибка при удалении дубликатов: {e}")
        return None


def build_order_payload(df, new_status, new_substatus):
    """
    Формирование списка заказов и ID
    """
    orders = []
    ids_sent = []

    for _, row in df.iterrows():
        try:
            order_id = int(row["HDRTAG2"])
            orders.append({
                "id": order_id,
                "status": new_status,
                "substatus": new_substatus
            })
            ids_sent.append(order_id)
        except Exception as e:
            logging.warning(f"Ошибка при обработке строки HDRTAG2={row.get('HDRTAG2')}: {e}")

    return orders, ids_sent


def send_status_update_request(payload, campaign_id, headers, max_retries=3, retry_delay=3):
    """
    Отправка запроса к API с повторами
    """
    url = f"https://api.partner.market.yandex.ru/campaigns/{campaign_id}/orders/status-update"
    logging.info(f"Отправка {len(payload['orders'])} заказов...")

    for attempt in range(1, max_retries + 1):
        try:
            response = requests.post(url, headers=headers, json=payload)

            if response.status_code == 200:
                logging.info("Статусы заказов успешно обновлены.")
                return True

            elif response.status_code in (429, 500, 502, 503, 504):
                logging.warning(f"Временная ошибка {response.status_code}, попытка {attempt}/{max_retries}")
                if attempt < max_retries:
                    time.sleep(retry_delay * attempt)
                    continue
                else:
                    logging.error("Превышено количество попыток.")
                    return False
            else:
                logging.error(f"Ошибка {response.status_code}: {response.text}")
                return False

        except requests.exceptions.RequestException as e:
            logging.exception(f"Сетевая ошибка, попытка {attempt}/{max_retries}: {e}")
            if attempt < max_retries:
                time.sleep(retry_delay * attempt)
                continue
            else:
                logging.error("Превышено количество попыток.")
                return False
        except Exception as e:
            logging.exception(f"Непредвиденная ошибка: {e}")
            return False


def update_order_statuses(df, campaign_id, headers, max_retries=3, retry_delay=3):
    """
    Обновляет статусы заказов в API Яндекс.Маркета.
    """
    if not validate_dataframe(df):
        return None

    unique_df = extract_unique_orders(df)
    if unique_df is None:
        return None

    orders, ids_sent = build_order_payload(unique_df, "PROCESSING", "READY_TO_SHIP")
    if not orders:
        logging.warning("Нет заказов для отправки.")
        return []

    payload = {"orders": orders}

    success = send_status_update_request(payload, campaign_id, headers, max_retries, retry_delay)
    return ids_sent if success else None



def process_confirmation_files(source_dir, processed_dir=CONFIRMATION_PROCESSED):
    """
    Обрабатывает .xls файлы из source_dir, объединяет в один DataFrame.
    Успешно обработанные файлы перемещает в processed_dir.
    """
    all_dataframes = []
    os.makedirs(processed_dir, exist_ok=True)

    for root, dirs, files in os.walk(source_dir):
        for file in files:
            if not file.endswith(".xls"):
                continue

            filepath = os.path.join(root, file)
            processed_path = os.path.join(processed_dir, file)

            if os.path.exists(processed_path):
                logging.info(f"Файл уже обработан ранее: {file}")
                continue

            try:
                df = append_data_from_file(filepath)
                all_dataframes.append(df)
                shutil.move(str(filepath), str(processed_path))
                logging.info(f"Файл перемещён в обработанные: {file}")
            except Exception as e:
                logging.error(f"Ошибка при обработке файла {file}: {e}")

    if all_dataframes:
        return pd.concat(all_dataframes, ignore_index=True)
    else:
        logging.warning("Не найдено подходящих файлов.")
        return pd.DataFrame()


def request_report_generation(order_ids, business_id, headers, format_pdf):
    url = "https://api.partner.market.yandex.ru/reports/documents/labels/generate"
    params = {'format': format_pdf}
    payload = {
        "businessId": business_id,
        "orderIds": order_ids,
        "sortingType": "SORT_BY_GIVEN_ORDER"
    }

    try:
        response = requests.post(url, headers=headers, params=params, json=payload)
        response.raise_for_status()
        data = response.json()
        logging.info("Запрос на генерацию ярлыков отправлен успешно.")
        logging.debug(f"Ответ: {json.dumps(data, indent=2, ensure_ascii=False)}")
        return data.get('result', {}).get('reportId')
    except requests.RequestException as e:
        logging.error(f"Ошибка при отправке запроса на генерацию: {e}")
        return None

def poll_report_status(report_id, headers, max_retries, retry_delay):
    url = f"https://api.partner.market.yandex.ru/reports/info/{report_id}"
    for attempt in range(max_retries):
        logging.info(f"Проверка статуса (попытка {attempt + 1}/{max_retries})...")
        try:
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            status_data = response.json()
            result = status_data.get('result', {})
            status = result.get('status')
            logging.info(f"Статус отчёта: {status}")
            if status == 'DONE':
                return result.get('file')
            elif status in ['PROCESSING', 'NEW', 'PENDING']:
                time.sleep(retry_delay)
            else:
                logging.error(f"Ошибка: статус {status}, подробности: {result.get('error', 'нет данных')}")
                return None
        except requests.RequestException as e:
            logging.error(f"Ошибка при проверке статуса: {e}")
            return None
    logging.error("Превышено количество попыток проверки статуса отчёта.")
    return None

def download_file(file_url, save_path, report_id, timeout, download_retries=3, retry_delay=5):
    filename = f"labels_{report_id}.pdf"
    file_path = os.path.join(save_path, filename)
    os.makedirs(save_path, exist_ok=True)

    for attempt in range(1, download_retries + 1):
        try:
            logging.info(f"Попытка скачивания {attempt}/{download_retries}: {file_url}")
            with requests.get(file_url, stream=True, timeout=timeout) as r:
                r.raise_for_status()
                with open(file_path, 'wb') as f:
                    for chunk in r.iter_content(chunk_size=8192):
                        f.write(chunk)
            logging.info(f"Файл успешно сохранён: {file_path}")
            return file_path
        except requests.exceptions.Timeout:
            logging.warning(f"Таймаут при скачивании (>{timeout} сек), попытка {attempt}.")
        except requests.RequestException as e:
            logging.warning(f"Ошибка при скачивании, попытка {attempt}: {e}")

        if attempt < download_retries:
            time.sleep(retry_delay)

    logging.error(f"Не удалось скачать файл после {download_retries} попыток.")
    return None


def generate_order_labels(order_ids, business_id, headers, format_pdf='A7',
                          max_retries=3, retry_delay=5, save_path=None,
                          download_timeout=15):
    report_id = request_report_generation(order_ids, business_id, headers, format_pdf)
    if not report_id:
        return None

    file_url = poll_report_status(report_id, headers, max_retries, retry_delay)
    if not file_url:
        return None

    if save_path:
        return download_file(file_url, save_path, report_id, download_timeout)

    else:
        logging.info("Генерация завершена. Возвращается ссылка на файл.")
        return file_url


def save_orders_to_excel(dataframe, output_path):
    """
    Сохраняет объединённый DataFrame в Excel-файл с заданной структурой:
    - Первый лист: список всех заказов
    - Второй лист: сводная таблица по позициям
    - Остальные листы: отдельные заказы

    :param dataframe: pd.DataFrame - Объединённый DataFrame со всеми заказами
    :param output_path: str - Путь для сохранения Excel-файла
    """

    if not os.path.exists(output_path):
        os.makedirs(output_path)

    id_recap = datetime.now().strftime("%d%m%y-%H%M%S")
    file_name = f'recap-YANDEX-{id_recap}.xlsx'
    path_file = os.path.join(output_path, file_name)

    with pd.ExcelWriter(path_file) as writer:
        # Первый лист: Список всех заказов
        all_orders = dataframe.pivot_table(
            index=['HDRTAG2', 'HDRTAG1'],
            values='QNT',
            aggfunc='sum'
        ).reset_index()
        all_orders.to_excel(writer, sheet_name='Список заказов', index=False)

        # Второй лист: Сводная таблица по позициям
        pivot_table = dataframe.pivot_table(
            index=['FIRM', 'CODEART', 'NAME', 'GDATE'],
            values='QNT',
            aggfunc='sum',
        ).reset_index()
        pivot_table.to_excel(writer, sheet_name='Сводная таблица',  index=False)

        #Третий лист: Сквозная таблица по позициям (лист подбора)
        dataframe.to_excel(writer, columns=['HDRTAG2', 'FIRM', 'CODEART', 'NAME', 'QNT'],
                           sheet_name="Лист подбора", index=False)

        # Остальные листы: Каждый заказ отдельно
        for order_id, order_data in dataframe.groupby('HDRTAG2'):
            order_data = order_data.drop(columns=['PODRCD', 'REFUSED', 'CODEPST'])
            order_data.to_excel(writer, sheet_name=f'ORDER {order_id}', index=False)

    return file_name



def set_common_properties(sheet: Worksheet, file_name: str, sfx):
    cm = 1 / 2.54
    sheet.page_margins = PageMargins(left=cm * 0.8, right=cm * 0.8, top=cm * 0.8, bottom=cm * 1.8)
    sheet.oddFooter.left.text = os.path.basename(file_name).split('.')[0] + " - " + sfx

    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(border_style='thin', color='FF000000'),
        right=Side(border_style='thin', color='FF000000'),
        top=Side(border_style='thin', color='FF000000'),
        bottom=Side(border_style='thin', color='FF000000'),
    )

    for row in range(1, sheet.max_row + 1):
        for column in range(1, sheet.max_column + 1):
            cell = sheet.cell(row, column)
            cell.alignment = alignment
            cell.border = thin_border
            cell.font = Font(name='Calibri', size=14)



def format_recap_sheet(sheet: Worksheet, file_name: str, sfx):
    sheet.insert_rows(1)
    sheet.cell(1, 1).font = Font(name='Calibri', size=20)
    sheet.cell(1, 1).value = os.path.basename(file_name).split('.')[0] + " - " + sfx
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 5
    sheet.print_title_rows = "2:2"



def format_merge_table_sheet(sheet: Worksheet, file_name: str, sfx):
    sheet.insert_rows(idx=1, amount=1)
    sheet.cell(1, 1).font = Font(name='Calibri', size=20)
    sheet.cell(1, 1).value = f"СВОДНАЯ ТАБЛИЦА К {os.path.basename(file_name).split('.')[0]} - {sfx}"
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 17
    sheet.column_dimensions['C'].width = 44
    sheet.column_dimensions['D'].width = 14
    sheet.column_dimensions['E'].width = 5
    sheet.print_title_rows = "2:2"

    # Инициализация стилей
    grey_fill = PatternFill(fill_type='solid', fgColor='FF808080')
    white_font = Font(color='ffffff', name='Calibri', size=14)
    thin_border = Border(bottom=Side(border_style='thin', color='ffffff'),
                         right=Side(border_style='thin', color='ffffff'))

    # Инициализация выравнивания и формата даты
    alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    date_format = 'DD.MM.YYYY'

    # Идем по строкам начиная с 3
    for row in range(3, sheet.max_row + 1):
        # Применяем выравнивание и форматирование для столбца 3 и 4
        sheet.cell(row, 3).alignment = alignment
        sheet.cell(row, 4).number_format = date_format

        # Форматирование ячейки 5, если значение больше 1
        if sheet.cell(row, 5).value > 1:
            sheet.cell(row, 5).fill = PatternFill(fill_type='solid', fgColor='FF000000')
            sheet.cell(row, 5).font = Font(color='ffffff', name='Calibri', size=14)
            sheet.cell(row, 5).border = Border(bottom=Side(border_style='thin', color='ffffff'))

        # Проверка повторяющихся значений в столбце 2
        current_value = sheet.cell(row, 2).value

        # Применяем форматирование только если это не первая ячейка, с которой начинается группа повторений
        if row < sheet.max_row and current_value == sheet.cell(row + 1, 2).value:
            # Если текущая ячейка повторяется с последующей, выделяем её серым
            sheet.cell(row, 2).fill = grey_fill
            sheet.cell(row, 2).font = white_font
            sheet.cell(row, 2).border = thin_border

            # Применяем серый фон и к следующей ячейке, если значения совпадают
            sheet.cell(row + 1, 2).fill = grey_fill
            sheet.cell(row + 1, 2).font = white_font
            sheet.cell(row + 1, 2).border = thin_border


            # Если текущая ячейка повторяется с последующей, выделяем её серым
            sheet.cell(row, 3).fill = grey_fill
            sheet.cell(row, 3).font = white_font
            sheet.cell(row, 3).border = thin_border

            # Применяем серый фон и к следующей ячейке, если значения совпадают
            sheet.cell(row + 1, 3).fill = grey_fill
            sheet.cell(row + 1, 3).font = white_font
            sheet.cell(row + 1, 3).border = thin_border

            # Если текущая ячейка повторяется с последующей, выделяем её серым
            sheet.cell(row, 4).fill = grey_fill
            sheet.cell(row, 4).font = white_font
            sheet.cell(row, 4).border = thin_border

            # Применяем серый фон и к следующей ячейке, если значения совпадают
            sheet.cell(row + 1, 4).fill = grey_fill
            sheet.cell(row + 1, 4).font = white_font
            sheet.cell(row + 1, 4).border = thin_border


        # Применяем выравнивание и форматирование для текущей строки
        sheet.cell(row, 3).alignment = alignment
        sheet.cell(row, 4).number_format = date_format



def format_orders_sheet(sheet: Worksheet):
    sheet.insert_rows(idx=1, amount=6)
    sheet.cell(1, 2).value = "НОМЕР ЗАКАЗА"
    sheet.cell(2, 2).value = "ДАТА ЗАКАЗА"
    sheet.cell(3, 2).value = "ДАТА ОТГРУЗКИ"
    sheet.cell(4, 2).value = "СТРОК В ЗАКАЗЕ"
    sheet.cell(5, 2).value = "ШТУК В ЗАКАЗЕ"
    sheet.cell(1, 4).value = sheet.cell(8, 1).value
    sheet.cell(2, 4).value = sheet.cell(8, 8).value
    sheet.cell(2, 4).number_format = 'DD.MM.YYYY'
    sheet.cell(3, 4).value = sheet.cell(8, 9).value
    sheet.cell(3, 4).number_format = 'DD.MM.YYYY'

    thin_border = Border(left=Side(border_style='thin', color='FF000000'),
                         right=Side(border_style='thin', color='FF000000'),
                         top=Side(border_style='thin', color='FF000000'),
                         bottom=Side(border_style='thin', color='FF000000'),
                         )

    alignment = Alignment(horizontal='center', vertical='center')
    font = Font(name='Calibri', size=16)

    sheet.delete_cols(1)
    sheet.delete_cols(7, 2)
    sheet.column_dimensions['A'].width = 14
    sheet.column_dimensions['B'].width = 16
    sheet.column_dimensions['C'].width = 36
    sheet.column_dimensions['D'].width = 4
    sheet.column_dimensions['E'].width = 13
    sheet.column_dimensions['F'].width = 14

    # Редактируем шапку файла
    # Проход по строкам
    for row in range(1, 6):
        # Объединение ячеек в строке
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        sheet.row_dimensions[row].height = 25

        # Проход по столбцам
        for col in range(1, 4):
            cell = sheet.cell(row=row, column=col)
            cell.alignment = alignment
            cell.border = thin_border
            cell.font = font

    # # Фарматируем тело таблицы
    stat_order = []  # для статистики по заказу
    for row in range(8, sheet.max_row + 1):
        sheet.cell(row, 3).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        sheet.cell(row, 5).alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        sheet.cell(row, 6).number_format = 'DD.MM.YYYY'
        stat_order.append(sheet.cell(row, 4).value)  # Собираем статистику
        if sheet.cell(row, 4).value > 1:
            sheet.cell(row, 4).fill = PatternFill(fill_type='solid', fgColor='FF000000')
            sheet.cell(row, 4).font = Font(color='ffffff', name='Calibri', size=14)
            sheet.cell(row, 4).border = Border(bottom=Side(border_style='thin', color='ffffff'))

    sheet.cell(4, 3).value = len(stat_order)
    sheet.cell(5, 3).value = sum(stat_order)
    sheet.oddFooter.right.text = f'Заказ {sheet.cell(1, 3).value}'
    sheet.print_options.horizontalCentered = True
    sheet.print_title_rows = "7:7"


def format_assembly_sheet(sheet: Worksheet, file_name: str, sfx):
    sheet.insert_rows(1)
    sheet.cell(1, 1).font = Font(name='Calibri', size=20)
    sheet.cell(1, 1).value = f"ЛИСТ ПОДБОРА К {os.path.basename(file_name).split('.')[0]} - {sfx}"
    sheet.page_setup.orientation = 'landscape'
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 18
    sheet.column_dimensions['C'].width = 18
    sheet.column_dimensions['D'].width = 67
    sheet.column_dimensions['E'].width = 4
    sheet.print_title_rows = "2:2"

    # Инициализация выравнивания и формата даты
    alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Инициализация стилей
    grey_fill = PatternFill(fill_type='solid', fgColor='FF808080')
    white_font = Font(color='ffffff', name='Calibri', size=14)
    thin_border = Border(bottom=Side(border_style='thin', color='ffffff'),
                         right=Side(border_style='thin', color='ffffff'))

    # Идем по строкам начиная с 3
    for row in range(3, sheet.max_row + 1):
        # Применяем выравнивание и форматирование для столбца 4
        sheet.cell(row, 4).alignment = alignment

        # Форматирование ячейки 5, если значение больше 1
        if sheet.cell(row, 5).value > 1:
            sheet.cell(row, 5).fill = PatternFill(fill_type='solid', fgColor='FF000000')
            sheet.cell(row, 5).font = Font(color='ffffff', name='Calibri', size=14)
            sheet.cell(row, 5).border = Border(bottom=Side(border_style='thin', color='ffffff'))

        # Проверка повторяющихся значений в столбце 2
        current_value = sheet.cell(row, 1).value

        # Применяем форматирование только если это не первая ячейка, с которой начинается группа повторений
        if row < sheet.max_row and current_value == sheet.cell(row + 1, 1).value:
            for col in range(1, 5):
                # Если текущая ячейка повторяется с последующей, выделяем её серым
                sheet.cell(row, col).fill = grey_fill
                sheet.cell(row, col).font = white_font
                sheet.cell(row, col).border = thin_border

                # Применяем серый фон и к следующей ячейке, если значения совпадают
                sheet.cell(row + 1, col).fill = grey_fill
                sheet.cell(row + 1, col).font = white_font
                sheet.cell(row + 1, col).border = thin_border



def format_report_sheets(report, suffix_shop):
    work_book = openpyxl.load_workbook(report)
    sheet_recap = work_book['Список заказов']
    sheet_pivot_table = work_book['Сводная таблица']
    sheet_assembly = work_book['Лист подбора']

    for sheet in work_book:
        set_common_properties(sheet, report, suffix_shop)

    format_recap_sheet(sheet_recap, report, suffix_shop)
    format_merge_table_sheet(sheet_pivot_table, report, suffix_shop)
    format_assembly_sheet(sheet_assembly, report, suffix_shop)

    for sheet in work_book.sheetnames[3:]:
        sheet_obj = work_book[sheet]
        format_orders_sheet(sheet_obj)

    work_book.save(report)


def main():
    # 1. Скачиваем подтверждения с FTP
    download_from_ftp(SERVER, USER_NAME, PASSWORD, CONFIRMATION_DIR, DIVISION_ID)

    # 2. Обрабатываем и объединяем данные из новых .xls файлов
    confirmations_df = process_confirmation_files(CONFIRMATION_DIR)

    # 3. Создаем отчет xlsx
    report_name = save_orders_to_excel(confirmations_df, RECAPS_DIR)
    format_report_sheets(os.path.join(RECAPS_DIR, report_name), 'FF')

    # 4. Получаем наклейки для заказов
    orders_id = confirmations_df['HDRTAG2'].unique().tolist()
    generate_order_labels(orders_id, BUSINESS_ID, HEADERS, save_path=STICKERS_DIR)

    # 5. Обновляем статусы заказов
    if not confirmations_df.empty:
        result = update_order_statuses(df=confirmations_df,
                                       campaign_id=CAMPAIGN_ID,
                                       headers=HEADERS)
        pprint(result)
    else:
        logging.info("Нет данных для обновления статусов.")

    # 6. Отправляем наклейки и рекап по email


if __name__ == "__main__":
    main()
